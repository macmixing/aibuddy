import os
import logging
import sys
import fitz  # PyMuPDF for PDF processing
import docx  # python-docx for DOCX processing
import openpyxl  # For XLSX processing
from striprtf.striprtf import rtf_to_text  # For RTF processing
import traceback

# Import configuration
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

def extract_text_from_pdf(file_path):
    """
    Extract text from PDF file
    
    Args:
        file_path (str): Path to the PDF file
        
    Returns:
        str: Extracted text
    """
    try:
        text = ""
        with fitz.open(file_path) as pdf:
            for page_num in range(len(pdf)):
                text += pdf[page_num].get_text()
        return text
    except Exception as e:
        logging.error(f"❌ Error extracting text from PDF: {e}")
        return f"Error extracting text from PDF: {str(e)}"

def extract_text_from_docx(file_path):
    """
    Extract text from DOCX file
    
    Args:
        file_path (str): Path to the DOCX file
        
    Returns:
        str: Extracted text
    """
    try:
        doc = docx.Document(file_path)
        text = "\n".join([paragraph.text for paragraph in doc.paragraphs if paragraph.text])
        return text
    except Exception as e:
        logging.error(f"❌ Error extracting text from DOCX: {e}")
        return f"Error extracting text from DOCX: {str(e)}"

def extract_text_from_xlsx(file_path):
    """
    Extract text from XLSX file
    
    Args:
        file_path (str): Path to the XLSX file
        
    Returns:
        str: Extracted text
    """
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        text = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text.append(f"\n--- Sheet: {sheet_name} ---\n")
            
            for row in sheet.iter_rows(values_only=True):
                # Filter out None values and convert to strings
                row_values = [str(cell) if cell is not None else "" for cell in row]
                # Only add non-empty rows
                if any(cell for cell in row_values):
                    text.append("\t".join(row_values))
        
        return "\n".join(text)
    except Exception as e:
        logging.error(f"❌ Error extracting text from XLSX: {e}")
        return f"Error extracting text from XLSX: {str(e)}"

def extract_text_from_rtf(file_path):
    """
    Extract text from RTF file
    
    Args:
        file_path (str): Path to the RTF file
        
    Returns:
        str: Extracted text
    """
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as rtf_file:
            rtf_content = rtf_file.read()
            text = rtf_to_text(rtf_content)
            return text
    except Exception as e:
        logging.error(f"❌ Error extracting text from RTF: {e}")
        return f"Error extracting text from RTF: {str(e)}"

def extract_text_from_txt(file_path):
    """
    Extract text from TXT file
    
    Args:
        file_path (str): Path to the TXT file
        
    Returns:
        str: Extracted text
    """
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as txt_file:
            return txt_file.read()
    except Exception as e:
        logging.error(f"❌ Error extracting text from TXT: {e}")
        return f"Error extracting text from TXT: {str(e)}"

def extract_text_from_file(file_path):
    """
    Extract text from a file based on its extension
    
    Args:
        file_path (str): Path to the file
        
    Returns:
        str: Extracted text
    """
    # Get file extension
    file_extension = os.path.splitext(file_path)[1].lower()
    
    # Extract text based on file extension
    if file_extension == '.pdf':
        return extract_text_from_pdf(file_path)
    elif file_extension == '.docx':
        return extract_text_from_docx(file_path)
    elif file_extension == '.xlsx':
        return extract_text_from_xlsx(file_path)
    elif file_extension == '.rtf':
        return extract_text_from_rtf(file_path)
    elif file_extension in ['.txt', '.csv', '.md', '.py', '.js', '.html', '.css', '.json', '.xml']:
        return extract_text_from_txt(file_path)
    else:
        return f"Unsupported file type: {file_extension}"